import os
import re
import io
import json
import base64
import logging
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from anthropic import Anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")
STORAGE_GROUP_ID = -5237650194

client = Anthropic(api_key=ANTHROPIC_API_KEY)

# In-memory: {invoice_name: [message_ids]} — one or more msgs per invoice
invoice_msg_ids = {}
# Cache: {invoice_name: inv_data} — avoid re-reading from group every time
invoice_cache = {}

# In-memory session
user_session = {"state": "idle", "active_invoice": None, "current_sku": None, "pending_products": []}

INVOICE_PROMPT = """You are an invoice parser. Extract all product lines from this invoice.
For each line extract SKU, product name, quantity, unit cost.
Respond ONLY with a JSON array:
[{"sku":"1308","name":"Dress Embroidery","qty":12,"cost":8.00}]
"""

COLOR_PARSE_PROMPT = """You parse clothing inventory entries. Extract colors, sizes and sale price.

OUTPUT: Respond with ONLY a JSON object. No explanation, no markdown, no code blocks.

JSON format:
{"colors":{"BLACK":{"code":"BLK","packs":3},"WHITE":{"code":"WHT","packs":3}},"sizes":["S","M","L"],"units_per_size":2,"sale_price":24.99}

RULES:
- sizes: extract size labels like S,M,L,XL,2XL,S/M,M/L
- units_per_size: number before each size (e.g. "2S2M2L" = 2)
- packs: number before each color
- sale_price: last number in the input (e.g. "24.99" or "24" at the end). If no price found, use 0.

COLOR MAPPING:
siyah/black=BLK, beyaz/white=WHT, mavi/blue=BLU, kirmizi/red=RED,
bej/beige=BGE, pembe/pink=PNK, yesil/green=GRN, gri/grey=GRY,
kahve/brown=BRN, turuncu/orange=ORG, mor/purple=PRP

EXAMPLES:
Input: "6 siyah 6 beyaz 2S2M2L 24.99"
Output: {"colors":{"BLACK":{"code":"BLK","packs":6},"WHITE":{"code":"WHT","packs":6}},"sizes":["S","M","L"],"units_per_size":2,"sale_price":24.99}

Input: "3siyah 3beyaz 2S2M2L 19"
Output: {"colors":{"BLACK":{"code":"BLK","packs":3},"WHITE":{"code":"WHT","packs":3}},"sizes":["S","M","L"],"units_per_size":2,"sale_price":19}

Input: "4 black 4 red 2S2M2L 15.50"
Output: {"colors":{"BLACK":{"code":"BLK","packs":4},"RED":{"code":"RED","packs":4}},"sizes":["S","M","L"],"units_per_size":2,"sale_price":15.50}
"""

# ── STORAGE ───────────────────────────────────────────────
# Each invoice saved as ONE message per invoice in the group.
# Message format: INV|invoice_name|{json_data}
# Max ~30 products per message. For larger invoices, split into INV1|, INV2|, etc.

def pack_products(products):
    """Minimize product data for storage"""
    result = []
    for p in products:
        s = {"s": p["sku"], "n": p["name"][:20], "q": p["qty"], "c": p["cost"], "d": int(p.get("completed", False))}
        if p.get("colors"): s["co"] = p["colors"]
        if p.get("sizes"): s["sz"] = p["sizes"]
        if p.get("units_per_size"): s["u"] = p["units_per_size"]
        if p.get("sale_price"): s["sp"] = p["sale_price"]
        if p.get("skipped"): s["sk"] = 1
        result.append(s)
    return result

def unpack_products(packed):
    result = []
    for p in packed:
        f = {"sku": p["s"], "name": p["n"], "qty": p["q"], "cost": p["c"], "completed": bool(p.get("d", 0))}
        if p.get("co"): f["colors"] = p["co"]
        if p.get("sz"): f["sizes"] = p["sz"]
        if p.get("u"): f["units_per_size"] = p["u"]
        if p.get("sp"): f["sale_price"] = p["sp"]
        if p.get("sk"): f["skipped"] = True
        result.append(f)
    return result

async def save_invoice(app, invoice_name, inv):
    """Save one invoice to group, splitting if needed"""
    global invoice_msg_ids
    products = pack_products(inv["products"])
    
    # Split into chunks of 20 to stay under 4096 chars
    chunks = [products[i:i+20] for i in range(0, len(products), 20)]
    
    # Delete old messages for this invoice
    if invoice_name in invoice_msg_ids:
        for msg_id in invoice_msg_ids[invoice_name] if isinstance(invoice_msg_ids[invoice_name], list) else [invoice_msg_ids[invoice_name]]:
            try:
                await app.bot.delete_message(STORAGE_GROUP_ID, msg_id)
            except:
                pass
    
    msg_ids = []
    total_chunks = len(chunks)
    for i, chunk in enumerate(chunks):
        text = f"INV|{invoice_name}|{i+1}/{total_chunks}|" + json.dumps(chunk, ensure_ascii=False, separators=(",",":"))
        try:
            msg = await app.bot.send_message(STORAGE_GROUP_ID, text)
            msg_ids.append(msg.message_id)
        except Exception as e:
            logger.error(f"Save chunk error: {e}")
    
    invoice_msg_ids[invoice_name] = msg_ids
    invoice_cache[invoice_name] = inv  # Update cache
    logger.info(f"Saved {invoice_name} in {len(msg_ids)} messages")

async def load_invoice(app, invoice_name):
    """Load invoice - use cache if available"""
    global invoice_cache
    if invoice_name in invoice_cache:
        return invoice_cache[invoice_name]
    
    if invoice_name not in invoice_msg_ids:
        return None
    
    msg_ids = invoice_msg_ids[invoice_name]
    if not isinstance(msg_ids, list):
        msg_ids = [msg_ids]
    
    chunks = {}
    for msg_id in msg_ids:
        try:
            fwd = await app.bot.forward_message(STORAGE_GROUP_ID, STORAGE_GROUP_ID, msg_id)
            if fwd.text and fwd.text.startswith("INV|"):
                parts = fwd.text.split("|", 3)
                if len(parts) == 4:
                    chunk_idx = int(parts[2].split("/")[0]) - 1
                    chunks[chunk_idx] = json.loads(parts[3])
            await app.bot.delete_message(STORAGE_GROUP_ID, fwd.message_id)
        except Exception as e:
            logger.error(f"Load chunk error: {e}")
    
    if not chunks:
        return None
    
    all_packed = []
    for i in sorted(chunks.keys()):
        all_packed.extend(chunks[i])
    
    inv = {"products": unpack_products(all_packed)}
    invoice_cache[invoice_name] = inv  # Cache it!
    return inv

async def load_all_invoices(app):
    """Load all invoices"""
    invoices = {}
    for name in list(invoice_msg_ids.keys()):
        inv = await load_invoice(app, name)
        if inv:
            invoices[name] = inv
    return invoices

async def delete_invoice(app, invoice_name):
    """Delete invoice messages from group"""
    global invoice_msg_ids, invoice_cache
    if invoice_name in invoice_msg_ids:
        msg_ids = invoice_msg_ids[invoice_name]
        if not isinstance(msg_ids, list):
            msg_ids = [msg_ids]
        for msg_id in msg_ids:
            try:
                await app.bot.delete_message(STORAGE_GROUP_ID, msg_id)
            except:
                pass
        del invoice_msg_ids[invoice_name]
    if invoice_name in invoice_cache:
        del invoice_cache[invoice_name]

# ── EXCEL ─────────────────────────────────────────────────

def create_excel(products, invoice_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"
    headers = [
        "Reference Handle","Token","Item Name","Customer-facing Name","Variation Name","SKU",
        "Description","Categories","Reporting Category","GTIN","Item Type","Weight (lb)",
        "Social Media Link Title","Social Media Link Description","Price","Online Sale Price",
        "Archived","Sellable","Contains Alcohol","Stockable","Skip Detail Screen in POS",
        "Option Name 1","Option Value 1","Default Unit Cost","Default Vendor Name",
        "Default Vendor Code","Current Quantity GAVA NEW YORK","New Quantity GAVA NEW YORK",
        "Stock Alert Enabled GAVA NEW YORK","Stock Alert Count GAVA NEW YORK"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", start_color="1F3864")
        cell.alignment = Alignment(horizontal="center")

    for p in products:
        description = f"{p['name']} — available in multiple colors and sizes."
        for color_name, color_info in p["colors"].items():
            qty_per_size = color_info["packs"] * p["units_per_size"]
            for size in p["sizes"]:
                size_slug = size.lower().replace("/", "-")
                handle = f"#{p['sku']}-{p['name'].lower().replace(' ','-')}-{color_name.lower()}-{size_slug}"
                sku_var = f"{p['sku']}-{color_info['code']}-{size}"
                row = {
                    "Reference Handle": handle, "Token": None,
                    "Item Name": p["name"], "Customer-facing Name": p["name"],
                    "Variation Name": f"{p['name']} {color_name} / {size}",
                    "SKU": sku_var, "Description": description,
                    "Categories": p.get("category",""), "Reporting Category": None, "GTIN": None,
                    "Item Type": "Physical good", "Weight (lb)": None,
                    "Social Media Link Title": None, "Social Media Link Description": None,
                    "Price": p.get("sale_price",0), "Online Sale Price": None, "Archived": "N",
                    "Sellable": None, "Contains Alcohol": "N", "Stockable": None,
                    "Skip Detail Screen in POS": "N", "Option Name 1": None, "Option Value 1": None,
                    "Default Unit Cost": p["cost"], "Default Vendor Name": p.get("vendor",""),
                    "Default Vendor Code": None,
                    "Current Quantity GAVA NEW YORK": qty_per_size,
                    "New Quantity GAVA NEW YORK": qty_per_size,
                    "Stock Alert Enabled GAVA NEW YORK": None,
                    "Stock Alert Count GAVA NEW YORK": None,
                }
                ws.append([row.get(h) for h in headers])

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def build_summary(products, invoice_name):
    lines = [f"✅ *{invoice_name} — Excel hazır!*\n"]
    total = 0
    for p in products:
        lines.append(f"📦 *{p['sku']} - {p['name']}*")
        for color_name, color_info in p["colors"].items():
            qty = color_info["packs"] * p["units_per_size"]
            for size in p["sizes"]:
                lines.append(f"  • {color_name}/{size} → {p['sku']}-{color_info['code']}-{size} | Qty:{qty} | ${p.get('sale_price',0)}")
                total += qty
        lines.append("")
    lines.append(f"📊 *Toplam: {total} adet*")
    return "\n".join(lines)

# ── INVOICE PARSING ───────────────────────────────────────

async def parse_invoice_image(image_bytes, media_type="image/jpeg"):
    image_b64 = base64.standard_b64encode(image_bytes).decode("utf-8")
    response = client.messages.create(
        model="claude-sonnet-4-20250514", max_tokens=2000,
        messages=[{"role":"user","content":[
            {"type":"image","source":{"type":"base64","media_type":media_type,"data":image_b64}},
            {"type":"text","text":INVOICE_PROMPT}
        ]}]
    )
    return response.content[0].text.strip()

async def parse_invoice_pdf(pdf_bytes):
    import fitz
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    all_products = []
    for page_num in range(len(doc)):
        pix = doc[page_num].get_pixmap(matrix=fitz.Matrix(2,2))
        result = await parse_invoice_image(pix.tobytes("jpeg"))
        json_match = re.search(r'\[[\s\S]*\]', result)
        if json_match:
            try: all_products.extend(json.loads(json_match.group()))
            except: pass
    doc.close()
    return json.dumps(all_products) if all_products else "[]"

# ── COMMANDS ──────────────────────────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_session.update({"state":"idle","active_invoice":None,"current_sku":None})
    await update.message.reply_text(
        "👋 Merhaba!\n\n"
        "📄 *Invoice yükle:* PDF veya fotoğraf gönder\n"
        "📋 *Bekleyen invoice'lar:* /invoices\n"
        "📦 *Koli geldiğinde:* `/teslim [isim]`\n"
        "✅ *Tüm kutular bitti:* /done\n"
        "⏭️ *Ürün atla:* `/skip SKU`",
        parse_mode="Markdown"
    )

async def cmd_invoices(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not invoice_msg_ids:
        await update.message.reply_text("📭 Kayıtlı invoice yok.")
        return
    invoices = await load_all_invoices(context.application)
    lines = ["📋 *Kayıtlı Invoice'lar:*\n"]
    for name, inv in invoices.items():
        total = len(inv["products"])
        done = sum(1 for p in inv["products"] if p.get("completed"))
        active = " ← AKTİF" if user_session.get("active_invoice") == name else ""
        lines.append(f"• *{name}*{active} — {done}/{total} tamamlandı")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")

async def cmd_teslim(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Kullanım: `/teslim vava mart`", parse_mode="Markdown")
        return
    invoice_name = " ".join(context.args).lower().strip()
    if invoice_name not in invoice_msg_ids:
        names = "\n".join(f"• {n}" for n in invoice_msg_ids.keys()) if invoice_msg_ids else "Hiç yok"
        await update.message.reply_text(f"❌ *{invoice_name}* bulunamadı.\n\nMevcut:\n{names}", parse_mode="Markdown")
        return
    inv = await load_invoice(context.application, invoice_name)
    if not inv:
        await update.message.reply_text("❌ Invoice yüklenemedi.")
        return
    user_session["active_invoice"] = invoice_name
    user_session["state"] = "receiving_products"
    remaining = [p for p in inv["products"] if not p.get("completed")]
    done_count = len(inv["products"]) - len(remaining)
    lines = [f"✅ *{invoice_name}* aktif!\n",
             f"📦 {len(remaining)} ürün bekliyor, {done_count} tamamlandı\n",
             "*Bekleyen ürünler:*"]
    for p in remaining[:20]:  # show max 20
        lines.append(f"• {p['sku']} - {p['name']} (${p['cost']})")
    if len(remaining) > 20:
        lines.append(f"• ... ve {len(remaining)-20} tane daha")
    lines.append("\n*Format:* `SKU renk/adet beden`")
    lines.append("_Örn: 1308 3siyah 3beyaz 2S2M2L_")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")

async def cmd_done(update: Update, context: ContextTypes.DEFAULT_TYPE):
    invoice_name = user_session.get("active_invoice")
    if not invoice_name:
        await update.message.reply_text("❌ Aktif invoice yok. `/teslim [isim]` yaz.", parse_mode="Markdown")
        return
    inv = await load_invoice(context.application, invoice_name)
    if not inv:
        await update.message.reply_text("❌ Invoice yüklenemedi.")
        return
    completed = [p for p in inv["products"] if p.get("completed")]
    remaining = [p for p in inv["products"] if not p.get("completed")]
    if not completed:
        await update.message.reply_text("❌ Henüz hiç ürün girilmedi.")
        return
    if remaining:
        names = ", ".join(p["sku"] for p in remaining[:10])
        await update.message.reply_text(
            f"⚠️ {len(remaining)} ürün girilmedi: *{names}*\n\nYine de devam: /done\_force",
            parse_mode="Markdown"
        )
        return
    await generate_excel(update, context, invoice_name, completed)

async def cmd_done_force(update: Update, context: ContextTypes.DEFAULT_TYPE):
    invoice_name = user_session.get("active_invoice")
    if not invoice_name:
        await update.message.reply_text("❌ Aktif invoice yok.")
        return
    inv = await load_invoice(context.application, invoice_name)
    if not inv:
        await update.message.reply_text("❌ Invoice yüklenemedi.")
        return
    completed = [p for p in inv["products"] if p.get("completed")]
    if not completed:
        await update.message.reply_text("❌ Hiç ürün girilmedi.")
        return
    await generate_excel(update, context, invoice_name, completed)

async def generate_excel(update, context, invoice_name, completed):
    await update.message.reply_text("⏳ Excel hazırlanıyor...")
    excel_buffer = create_excel(completed, invoice_name)
    summary = build_summary(completed, invoice_name)
    await update.message.reply_text(summary, parse_mode="Markdown")
    safe_name = invoice_name.replace(" ", "_")
    filename = f"{safe_name}_square_catalog.xlsx"
    await update.message.reply_document(
        document=excel_buffer, filename=filename,
        caption=f"📎 `{filename}`", parse_mode="Markdown"
    )
    await delete_invoice(context.application, invoice_name)
    user_session.update({"active_invoice": None, "state": "idle", "current_sku": None})

async def cmd_skip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Kullanım: `/skip 1308`", parse_mode="Markdown")
        return
    invoice_name = user_session.get("active_invoice")
    if not invoice_name:
        await update.message.reply_text("❌ Aktif invoice yok.")
        return
    sku = context.args[0].upper()
    inv = await load_invoice(context.application, invoice_name)
    if not inv:
        await update.message.reply_text("❌ Invoice yüklenemedi.")
        return
    for p in inv["products"]:
        if p["sku"].upper() == sku:
            p["completed"] = True
            p["skipped"] = True
            await save_invoice(context.application, invoice_name, inv)
            remaining = sum(1 for x in inv["products"] if not x.get("completed"))
            await update.message.reply_text(f"⏭️ *{sku}* atlandı. {remaining} ürün kaldı.", parse_mode="Markdown")
            return
    await update.message.reply_text(f"❌ {sku} bulunamadı.")

async def cmd_reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_session.update({"state":"idle","active_invoice":None,"current_sku":None})
    await update.message.reply_text("🔄 Sıfırlandı!")

# ── FILE HANDLERS ─────────────────────────────────────────

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.document.mime_type != "application/pdf":
        await update.message.reply_text("❌ Sadece PDF veya fotoğraf.")
        return
    await update.message.reply_text("📄 Invoice okunuyor, lütfen bekle...")
    try:
        file = await context.bot.get_file(update.message.document.file_id)
        pdf_bytes = await file.download_as_bytearray()
        result = await parse_invoice_pdf(bytes(pdf_bytes))
        await process_invoice_result(update, context, result)
    except Exception as e:
        logger.error(f"PDF error: {e}")
        await update.message.reply_text(f"❌ PDF okunamadı: {str(e)}")

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("📸 Invoice okunuyor...")
    try:
        photo = update.message.photo[-1]
        file = await context.bot.get_file(photo.file_id)
        img_bytes = await file.download_as_bytearray()
        result = await parse_invoice_image(bytes(img_bytes))
        await process_invoice_result(update, context, result)
    except Exception as e:
        logger.error(f"Photo error: {e}")
        await update.message.reply_text(f"❌ Fotoğraf okunamadı: {str(e)}")

async def process_invoice_result(update, context, result):
    try:
        json_match = re.search(r'\[[\s\S]*\]', result)
        if not json_match:
            await update.message.reply_text("❌ Invoice okunamadı.")
            return
        products = json.loads(json_match.group())
        for p in products:
            p["completed"] = False
        lines = [f"✅ *{len(products)} ürün tespit edildi:*\n"]
        for p in products[:15]:
            lines.append(f"• {p['sku']} - {p['name']} | {p['qty']} adet | ${p['cost']}")
        if len(products) > 15:
            lines.append(f"• ... ve {len(products)-15} tane daha")
        lines.append("\n*Bu invoice'a bir isim ver:*\n_(örn: vava mart, supplier abc)_")
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
        user_session["state"] = "waiting_invoice_name"
        user_session["pending_products"] = products
    except Exception as e:
        logger.error(f"Invoice processing error: {e}")
        await update.message.reply_text("❌ Hata oluştu.")

# ── MESSAGE HANDLER ───────────────────────────────────────

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_message = update.message.text.strip()
    state = user_session.get("state", "idle")

    if state == "waiting_invoice_name":
        invoice_name = user_message.lower().strip()
        if invoice_name in invoice_msg_ids:
            await update.message.reply_text(f"⚠️ *{invoice_name}* zaten var. Farklı isim yaz.", parse_mode="Markdown")
            return
        products = user_session.get("pending_products", [])
        await save_invoice(context.application, invoice_name, {"products": products})
        user_session["state"] = "idle"
        user_session["pending_products"] = []
        await update.message.reply_text(
            f"✅ *{invoice_name}* kaydedildi! {len(products)} ürün bekleniyor.\n\nKoli geldiğinde:\n`/teslim {invoice_name}`",
            parse_mode="Markdown"
        )
        return

    if state == "receiving_products":
        invoice_name = user_session.get("active_invoice")
        if not invoice_name:
            await update.message.reply_text("❌ Aktif invoice yok.")
            return
        inv = await load_invoice(context.application, invoice_name)
        if not inv:
            await update.message.reply_text("❌ Invoice yüklenemedi.")
            return

        # Support multiple products per message (one per line)
        lines = [l.strip() for l in user_message.strip().split("\n") if l.strip()]
        saved = []
        errors = []

        for line in lines:
            parts = line.split()
            if not parts: continue
            sku = parts[0].upper()
            rest = " ".join(parts[1:])

            product = next((p for p in inv["products"] if p["sku"].upper() == sku), None)
            if not product:
                errors.append(f"❌ *{sku}* bu invoice'da yok")
                continue
            if product.get("completed"):
                errors.append(f"⚠️ *{sku}* zaten girilmiş")
                continue
            try:
                response = client.messages.create(
                    model="claude-sonnet-4-20250514", max_tokens=500,
                    system=COLOR_PARSE_PROMPT,
                    messages=[{"role":"user","content":rest}]
                )
                raw = response.content[0].text.strip()
                # Extract JSON even if Claude adds extra text
                json_match = re.search(r'\{[\s\S]*\}', raw)
                if not json_match:
                    errors.append(f"❌ *{sku}* parse edilemedi")
                    logger.error(f"No JSON in response for {sku}: {raw}")
                    continue
                parsed = json.loads(json_match.group())
                product["colors"] = parsed["colors"]
                product["sizes"] = parsed["sizes"]
                product["units_per_size"] = parsed["units_per_size"]
                product["sale_price"] = parsed.get("sale_price", 0)
                product["completed"] = True
                saved.append(sku)
            except Exception as e:
                logger.error(f"Parse error {sku}: {e}")
                errors.append(f"❌ *{sku}* anlaşılamadı: {str(e)[:50]}")

        if saved:
            # Update cache immediately - group save happens in background
            invoice_cache[invoice_name] = inv
            # Save to group (don't await - let it run)
            import asyncio
            asyncio.create_task(save_invoice(context.application, invoice_name, inv))

        remaining = [p for p in inv["products"] if not p.get("completed")]
        msg_lines = []
        if saved:
            msg_lines.append(f"✅ *{', '.join(saved)}* kaydedildi!")
        if errors:
            msg_lines.extend(errors)
        msg_lines.append(f"\n📦 *{len(remaining)} ürün kaldı.*")
        if not remaining:
            msg_lines.append("\n🎉 Tüm ürünler bitti! /done yaz.")

        await update.message.reply_text("\n".join(msg_lines), parse_mode="Markdown")
        return

    await update.message.reply_text(
        "📄 Invoice için PDF veya fotoğraf yükle.\n📋 Mevcut invoice'lar için /invoices",
        parse_mode="Markdown"
    )

def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("invoices", cmd_invoices))
    app.add_handler(CommandHandler("teslim", cmd_teslim))
    app.add_handler(CommandHandler("done", cmd_done))
    app.add_handler(CommandHandler("done_force", cmd_done_force))
    app.add_handler(CommandHandler("skip", cmd_skip))
    app.add_handler(CommandHandler("reset", cmd_reset))
    app.add_handler(MessageHandler(filters.Document.PDF, handle_document))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    logger.info("Bot starting...")
    app.run_polling(allowed_updates=Update.ALL_TYPES, drop_pending_updates=True)

if __name__ == "__main__":
    main()
